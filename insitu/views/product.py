# -*- coding: utf-8 -*-
from django.contrib import messages
from django.core.urlresolvers import reverse
from django.db import transaction
from django.db.models.fields.reverse_related import ForeignObjectRel
from django.http.response import HttpResponse
from django.utils.safestring import mark_safe
from django.template import Library
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from insitu import documents
from insitu import forms
from insitu import models
from insitu.signals import DisableSignals
from insitu.utils import get_choices
from insitu.views.base import ESDatatableView
from insitu.views.protected import (
    ProtectedView,
    ProtectedTemplateView, ProtectedDetailView,
    ProtectedUpdateView, ProtectedCreateView,
    ProtectedDeleteView,
    LoggingProtectedCreateView,
    )
from insitu.views.protected.permissions import (
    IsAuthenticated,
    IsSuperuser,
)
from picklists import models as pickmodels
from picklists.views import solve_sql

import json

class ProductList(ProtectedTemplateView):
    template_name = 'product/list.html'
    permission_classes = (IsAuthenticated, )
    permission_denied_redirect = None
    target_type = 'products'

    def permission_denied(self, request):
        self.permission_denied_redirect = reverse('auth:login')
        return super().permission_denied(request)

    def get_context_data(self):
        context = super(ProductList, self).get_context_data()
        services = get_choices('name', model_cls=models.CopernicusService)
        groups = get_choices('name', model_cls=pickmodels.ProductGroup)
        statuses = get_choices('name', model_cls=pickmodels.ProductStatus)
        areas = get_choices('name', model_cls=pickmodels.Area)
        components = get_choices('name', model_cls=models.Component)
        entities = get_choices('acronym', model_cls=models.EntrustedEntity)
        context.update({
            'services': services,
            'groups': groups,
            'statuses': statuses,
            'areas': areas,
            'components': components,
            'entities': entities,
        })
        return context


class ProductListJson(ESDatatableView):
    columns = ['name', 'service', 'entity', 'component', 'group',
               'status', 'area']
    order_columns = columns
    filters = ['service', 'entity', 'component', 'group', 'status', 'area']
    filter_fields = [
        'component__service__name', 'component__entrusted_entity__acronym',
        'component__name', 'group__name', 'status__name', 'area__name',
    ]  # This must be in the same order as `filters`
    document = documents.ProductDoc
    permission_classes = (IsAuthenticated, )


class ProductAdd(ProtectedCreateView):
    template_name = 'product/add.html'
    form_class = forms.ProductForm
    permission_classes = (IsSuperuser, )
    target_type = 'product'

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'The product was created successfully!')
        return response

    def permission_denied(self, request):
        self.permission_denied_redirect = reverse('product:list')
        return super().permission_denied(request)

    def get_success_url(self):
        return reverse('product:list')


class ProductEdit(ProtectedUpdateView):
    template_name = 'product/edit.html'
    form_class = forms.ProductForm
    model = models.Product
    context_object_name = 'product'
    permission_classes = (IsSuperuser, )
    target_type = 'product'

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'The product was updated successfully!')
        return response

    def permission_denied(self, request):
        self.permission_denied_redirect = reverse('product:list')
        return super().permission_denied(request)

    def get_success_url(self):
        product = self.get_object()
        return reverse('product:detail', kwargs={'pk': product.pk})


class ProductDetail(ProtectedDetailView):
    template_name = 'product/detail.html'
    model = models.Product
    context_object_name = 'product'
    permission_classes = (IsAuthenticated, )
    target_type = 'product'

    def permission_denied(self, request):
        self.permission_denied_redirect = reverse('auth:login')
        return super().permission_denied(request)


class ProductDelete(ProtectedDeleteView):

    template_name = 'product/delete.html'
    form_class = forms.ProductForm
    model = models.Product
    context_object_name = 'product'
    permission_classes = (IsSuperuser, )
    target_type = 'product'

    def permission_denied(self, request):
        self.permission_denied_redirect = reverse('product:list')
        return super().permission_denied(request)

    def get_success_url(self):
        messages.success(self.request, 'The product was deleted successfully!')
        return reverse('product:list')


SKIP_FIELDS = ['_deleted', 'requirements', 'created_at', 'updated_at']


class ExportProductView(ProtectedView):
    permission_classes = (IsSuperuser,)

    def get(self, request, *args, **kwargs):
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="products.xlsx"'
        wb = Workbook()
        wb.remove(wb.active)

        ws = wb.create_sheet('product')
        # Header
        current_row = 1
        columns = [field for field in models.Product._meta.get_fields()
                   if not isinstance(field, ForeignObjectRel) and field.name not in
                   SKIP_FIELDS]
        for col in columns:
            cell = ws.cell(row=current_row,
                           column=columns.index(col) + 1,
                           value=col.name)
            cell.font = Font(bold=True)

        # Data
        data = models.Product.objects.all()
        for entry in data:
            current_row += 1
            for col in columns:
                value = getattr(entry, col.name)
                if col.related_model:
                    value = value.pk
                ws.cell(row=current_row,
                        column=columns.index(col) + 1,
                        value=value)
        ws.freeze_panes = 'A2'
        wb.save(response)
        return response


RELATED_FIELDS = {
    'group': pickmodels.ProductGroup,
    'component': models.Component,
    'status': pickmodels.ProductStatus,
    'area': pickmodels.Area,
}


class ImportProductsView(ProtectedView):
    permission_classes = (IsSuperuser,)

    def post(self, request, *args, **kwargs):
        if 'workbook' not in request.FILES:
            return HttpResponse(status=400)
        workbook_file = request.FILES['workbook']
        try:
            wb = load_workbook(workbook_file)
            with transaction.atomic():
                ws = wb.get_sheet_by_name('product')
                fields = [
                    col[0].value
                    for col in ws.iter_cols(min_row=1, max_row=1)
                    if col[0].value is not None
                ]
                with DisableSignals():
                    for row in ws.iter_rows(min_row=2):
                        data = {}
                        pk = row[0].value
                        for i in range(1, len(fields)):
                            field = fields[i]
                            value = row[i].value if row[i].value is not None else ''
                            if field in RELATED_FIELDS.keys():
                                value = RELATED_FIELDS[field].objects.get(pk=value)
                            data[field] = value
                        if not [val for val in data.values() if val]:
                            continue
                        data['_deleted'] = False
                        models.Product.objects.really_all().update_or_create(
                            id=pk,
                            defaults=data
                        )

            solve_sql()
            from django.core.management import call_command
            call_command('search_index', '--rebuild', '-f')
        except Exception:
            return HttpResponse(status=400)
        return HttpResponse(status=200)


register = Library()

@register.filter(is_safe=True)
def js(obj):
    return mark_safe(json.dumps(obj))
