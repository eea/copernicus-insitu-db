import datetime
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook

from django.urls import reverse_lazy
from django.conf import settings
from django.shortcuts import get_object_or_404

from insitu.views import protected
from insitu.views.protected.views import ProtectedTemplateView
from insitu.utils import PICKLISTS_DESCRIPTION
from picklists import models

from explorer.exporters import get_exporter_class
from explorer.forms import QueryForm
from explorer.models import Query
from explorer.views import (
    DownloadQueryView,
    PlayQueryView,
    _export,
    query_viewmodel
)
from explorer.utils import extract_params
from explorer.utils import (
    url_get_rows,
)

class Manager(ProtectedTemplateView):
    template_name = 'manage.html'
    permission_classes = (protected.IsSuperuser,)
    permission_denied_redirect = reverse_lazy('auth:login')


class HelpPage(ProtectedTemplateView):
    template_name = 'help.html'
    permission_classes = (protected.IsAuthenticated,)
    permission_denied_redirect = reverse_lazy('auth:login')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['models'] = dict()

        PICKLISTS = [
            models.Barrier, models.ComplianceLevel, models.Area,
            models.Criticality, models.Country, models.DataFormat,
            models.DataPolicy, models.DataType, models.DefinitionLevel,
            models.Dissemination, models.EssentialVariable, models.InspireTheme,
            models.ProductGroup, models.ProductStatus, models.ProviderType,
            models.Relevance, models.RequirementGroup,
            models.QualityControlProcedure, models.Timeliness,
            models.UpdateFrequency,
        ]

        for model in PICKLISTS:
            data = {
                'nice_name': model._meta.verbose_name,
                'description': PICKLISTS_DESCRIPTION.get(model.__name__, None),
                'objects': model.objects.order_by('pk'),
                'fields': [field.name for field in model._meta.fields
                           if field.name not in ('id', 'sort_order')]
            }
            context['models'][model._meta.model_name] = data
            context['email'] = settings.SUPPORT_EMAIL
        return context


class AboutView(ProtectedTemplateView):
    template_name = 'about.html'
    permission_classes = (protected.IsAuthenticated,)
    permission_denied_redirect = reverse_lazy('auth:login')


class ReportsListView(ProtectedTemplateView):
    template_name = 'reports/list.html'
    permission_classes = (protected.IsAuthenticated,)
    permission_denied_redirect = reverse_lazy('auth:login')

    def get_context_data(self, **kwargs):
        context = super(ReportsListView, self).get_context_data(**kwargs)
        context['queries'] = Query.objects.all().order_by('id').values(
            'id', 'title', 'description')
        return context


class ReportsDetailView(ProtectedTemplateView):
    template_name = 'reports/detail.html'
    permission_classes = (protected.IsAuthenticated,)
    permission_denied_redirect = reverse_lazy('auth:login')

    def get(self, request, *args, **kwargs):
        self.report = get_object_or_404(Query, pk=kwargs['query_id'])
        return super(ReportsDetailView, self).get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super(ReportsDetailView, self).get_context_data(**kwargs)
        context['query'] = {
            'id': self.report.id,
            'title': self.report.title,
            'description': self.report.description,
            'params': extract_params(self.report.sql),
        }
        return context


def as_text(value):
    if value is None:
        return ""
    return str(value)


class PlaygroundView(PlayQueryView):

    def render(self):
        return self.render_template(
            'reports/playground.html',
            {'title': 'Playground', 'form': QueryForm(),
             'no_jquery': True}
        )

    def render_with_sql(self, request, query, run_query=True, error=None):
        rows = url_get_rows(request)
        context = query_viewmodel(request.user, query, title="Playground",
                            run_query=run_query,
                            error=error, rows=rows)
        context.update({'no_jquery': True})
        return self.render_template('reports/playground.html', context)


class DownloadReportView(DownloadQueryView):

    def get(self, request, query_id, *args, **kwargs):
        query = get_object_or_404(Query, pk=query_id)
        format = request.GET.get('format', 'csv')
        exporter_class = get_exporter_class(format)
        file_extension = exporter_class.file_extension
        date = '_' + datetime.datetime.now().strftime('%Y%m%d')
        response = _export(request, query)
        response['Content-Disposition'] = (date + file_extension).join(
            response['Content-Disposition'].split(file_extension))
        if file_extension == '.xlsx':
            wb = load_workbook(filename=BytesIO(response.content))
            ws = wb.active
            dims = {}
            for row in ws.iter_rows():
                for cell in row:
                    dims[cell.column] = max(dims.get(cell.column, 0), len(as_text(cell.value)))
                    # cell.alignment = cell.alignment.copy(wrapText=True)
            for col, value in dims.items():
                ws.column_dimensions[col].width = value
            wb.close()
            virtaul_wb = save_virtual_workbook(wb)
            response.content = virtaul_wb
        return response
