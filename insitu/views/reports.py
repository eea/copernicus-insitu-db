import string

from datetime import datetime
from io import BytesIO
from openpyxl import load_workbook

from explorer.app_settings import UNSAFE_RENDERING
from explorer.exporters import get_exporter_class
from explorer.models import Query
from explorer.views import DownloadQueryView
from explorer.views.export import _export
from explorer.utils import extract_params
from wkhtmltopdf.views import PDFTemplateResponse

from django.conf import settings
from django.http import HttpResponse, JsonResponse, Http404
from django.shortcuts import get_object_or_404
from django.template.loader import get_template
from django.urls import reverse_lazy
from django.views import View

from insitu.forms import (
    CountryReportForm,
    DataNetworkReportForm,
    StandardReportForm,
    DataProviderDuplicatesReportForm,
    EntriesCountReportForm,
    UserActionsForm,
)
from insitu.models import (
    Component,
    CopernicusService,
    Product,
    DataProvider,
)
from insitu.views._reports import (
    CountryReportExcelMixin,
    CountryReportPDFMixin,
    DataProviderDuplicatesReportMixin,
    DataProviderNetworkReportExcelMixin,
    EntriesCountReportExcelMixin,
    UserActionsReportMixin,
    StandardReportExcelMixin,
    StandardReportPDFMixin,
)
from insitu.views.protected.permissions import IsAuthenticated
from insitu.views.protected.views import ProtectedTemplateView, ProtectedView
from insitu.utils import as_text
from picklists.models import Country


class ReportsListView(ProtectedTemplateView):
    template_name = "reports/list.html"
    permission_classes = ()
    permission_denied_redirect = reverse_lazy("auth:login")

    def get_context_data(self, **kwargs):
        context = super(ReportsListView, self).get_context_data(**kwargs)
        if self.request.user.is_authenticated:
            context["queries"] = (
                Query.objects.exclude(id__in=settings.EXCLUDE_REPORTS_IDS)
                .order_by("id")
                .values("id", "title", "description")
            )
        else:
            context["queries"] = (
                Query.objects.filter(id__in=settings.PUBLIC_REPORTS_IDS)
                .order_by("id")
                .values("id", "title", "description")
            )
        return context


class ReportsDetailView(ProtectedTemplateView):
    template_name = "reports/detail.html"
    permission_classes = ()
    permission_denied_redirect = reverse_lazy("auth:login")

    def get(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            if kwargs["query_id"] not in settings.PUBLIC_REPORTS_IDS:
                return HttpResponse(
                    "You don't have permission to access this page", status=403
                )
        try:
            pk = int(kwargs["query_id"])
            self.report = get_object_or_404(Query, pk=pk)
            return super(ReportsDetailView, self).get(request, *args, **kwargs)
        except ValueError:
            return HttpResponse("Invalid query id", status=400)

    def get_filename(self, title):
        valid_chars = "-_.() %s%s" % (string.ascii_letters, string.digits)
        filename = "".join(c for c in title if c in valid_chars)
        filename = filename.replace(" ", "_")
        return filename

    def get_context_data(self, **kwargs):
        context = super(ReportsDetailView, self).get_context_data(**kwargs)
        self.report.execute_query_only()
        context["query"] = {
            "id": self.report.id,
            "title": self.report.title,
            "description": self.report.description,
            "params": extract_params(self.report.sql),
        }
        filename = self.get_filename(self.report.title) + datetime.now().strftime(
            "%Y%m%d"
        )
        context.update(
            {
                "html_filename": filename + ".html",
                "pdf_filename": filename + ".pdf",
                "excel_filename": filename + ".xlsx",
                "no_jquery": True,
                "unsafe_rendering": UNSAFE_RENDERING,
            }
        )
        return context


class ReportDataJsonView(ProtectedView):
    def get(self, request, *args, **kwargs):
        self.report = get_object_or_404(Query, pk=kwargs["query_id"])
        res = self.report.execute_query_only()
        data = []
        for row in res.data:
            column_count = 0
            row_data = {}
            for column in res.headers:
                row_data[column.title] = str(row[column_count])
                column_count = column_count + 1
            data.append(row_data)

        return JsonResponse(data, safe=False)


class SnapshotView(ProtectedTemplateView):
    def get(self, request, *args, **kwargs):
        response = HttpResponse()
        date = datetime.now().strftime("%Y%m%d")
        response["Content-Disposition"] = (
            "attachment; filename=insitu_{0}.sql.gz".format(date)
        )
        response["X-Accel-Redirect"] = "/static/protected/database.sql.gz"
        return response


class DownloadReportsView(DownloadQueryView):
    def get(self, request, query_id, *args, **kwargs):
        query = get_object_or_404(Query, pk=query_id)
        format = request.GET.get("format", "csv")
        exporter_class = get_exporter_class(format)
        file_extension = exporter_class.file_extension
        date = "_" + datetime.now().strftime("%Y%m%d")
        response = _export(request, query)
        response["Content-Disposition"] = (date + file_extension).join(
            response["Content-Disposition"].split(file_extension)
        )
        if file_extension == ".xlsx":
            wb = load_workbook(filename=BytesIO(response.content))
            ws = wb.active
            dims = {}
            for row in ws.iter_rows():
                for cell in row:
                    dims[cell.column] = max(
                        dims.get(cell.column, 0), len(as_text(cell.value))
                    )
            for col, value in dims.items():
                ws.column_dimensions[col].width = value
            wb.close()
            with BytesIO() as buffer:
                wb.save(buffer)
            response.content = buffer.getvalue()
        return response


class HTMLToPDFView(View):
    def render(self, context, request):
        template = get_template("reports/reports_pdf.html")
        template_response = PDFTemplateResponse(
            request=request,
            template=template,
            filename="test.pdf",
            context=context,
            show_content_in_browser=False,
            cmd_options={
                "javascript-delay": 3000,
            },
        )
        template_response.render()
        content = template_response.rendered_content
        return HttpResponse(content, content_type="application/pdf")

    def post(self, request, *args, **kwargs):
        context = {
            "data": request.POST["data"],
            "title": request.POST.get("title", ""),
            "date": datetime.now().strftime("%Y %m %d"),
        }
        return self.render(context, request)


class StandardReportView(
    ProtectedTemplateView, StandardReportExcelMixin, StandardReportPDFMixin
):
    template_name = "reports/standard_report.html"
    permission_classes = ()
    permission_denied_redirect = reverse_lazy("auth:login")

    def get_filtered_services(self):
        # only take those services into consideration
        services = CopernicusService.objects.filter(
            acronym__in=["CEMS", "CLMS", "CSS", "CAMS", "C3S", "CMEMS", "CSC"]
        )
        return services

    def get_filtered_components(self, services):
        components = Component.objects.filter(service__in=services).exclude(
            acronym__in=["MWR", "SRAL", "OLCI", "SLSTR", "SLSTR/OLCI", "TROPOMI"]
        )
        return components

    def get_context_data(self, **kwargs):
        context = super(StandardReportView, self).get_context_data(**kwargs)
        context["services"] = self.get_filtered_services()
        context["components"] = self.get_filtered_components(context["services"])
        context["form"] = StandardReportForm()
        return context

    def post(self, request, *args, **kwargs):
        services = self.request.POST.getlist("service")
        for service in services:
            try:
                int(service)
            except ValueError:
                raise Http404("Invalid service selected")
        components = self.request.POST.getlist("component")
        for component in components:
            try:
                int(component)
            except ValueError:
                raise Http404("Invalid component selected")
        self.services = CopernicusService.objects.filter(id__in=services)
        self.components = Component.objects.filter(id__in=components)
        self.products = Product.objects.filter(component_id__in=components).order_by(
            "name"
        )
        action = request.POST.get("action", "")
        services = "_".join([service.acronym for service in self.services])
        components = "_".join([component.acronym for component in self.components])
        base_filename = "_".join(["Standard_Report", services, components])
        if action == "Generate PDF":
            filename = self.generate_filename(base_filename, "pdf")
            return self.generate_pdf(filename)
        elif action == "Generate Excel":
            filename = self.generate_filename(base_filename, "xlsx")
            return self.generate_excel(filename)
        else:
            return HttpResponse("Incorect value selected")


class CountryReportView(
    ProtectedTemplateView, CountryReportExcelMixin, CountryReportPDFMixin
):
    template_name = "reports/country_report.html"
    permission_classes = ()
    permission_denied_redirect = reverse_lazy("auth:login")

    def get_context_data(self, **kwargs):
        context = super(CountryReportView, self).get_context_data(**kwargs)
        context["countries"] = Country.objects.all()
        context["form"] = CountryReportForm()
        return context

    def post(self, request, *args, **kwargs):
        self.country_code = self.request.POST.getlist("country")[0]
        self.dataproviders = DataProvider.objects.filter(
            countries__code=self.country_code, is_network=False, _deleted=False
        ).order_by("name")
        self.dataproviders_networks = DataProvider.objects.filter(
            countries__code=self.country_code, is_network=True, _deleted=False
        ).order_by("name")
        country = Country.objects.get(code=self.country_code).name
        base_filename = "_".join([country, "Country_Report"])
        action = request.POST.get("action", "")
        if action == "Generate PDF":
            filename = self.generate_filename(base_filename, "pdf")
            return self.generate_pdf(filename)
        elif action == "Generate Excel":
            filename = self.generate_filename(base_filename, "xlsx")
            return self.generate_excel(filename)
        else:
            return HttpResponse("Incorect value selected")


class DataProviderDuplicatesReportView(
    ProtectedTemplateView, DataProviderDuplicatesReportMixin
):
    template_name = "reports/data_provider_duplicates_report.html"
    permission_classes = (IsAuthenticated,)
    permission_denied_redirect = reverse_lazy("auth:login")

    def get_context_data(self, **kwargs):
        context = super(DataProviderDuplicatesReportView, self).get_context_data(
            **kwargs
        )
        context["countries"] = Country.objects.all()
        context["form"] = DataProviderDuplicatesReportForm()
        return context

    def post(self, request, *args, **kwargs):
        filename = self.generate_filename(
            "CIS2_Potential_Provider_Duplicates_Report", "xlsx"
        )
        return self.generate_excel(filename)


class DataProvidersNetworkReportView(
    ProtectedTemplateView, DataProviderNetworkReportExcelMixin
):
    template_name = "reports/data_provider_network_report.html"
    permission_classes = (IsAuthenticated,)
    permission_denied_redirect = reverse_lazy("auth:login")

    def get_context_data(self, **kwargs):
        context = super(DataProvidersNetworkReportView, self).get_context_data(
            **kwargs
        )
        context["form"] = DataNetworkReportForm()
        return context

    def post(self, request, *args, **kwargs):
        # TODO: Move this hardcoded values to a boolean on the model
        self.ACCEPTED_NETWORKS_IDS = [
            802,
            21,
            122,
            23,
            134,
            811,
            828,
            358,
            839,
            2,
            827,
            180,
            890,
            16,
            796,
            602,
            600,
        ]

        self.ACCEPTED_RESEARCH_INFRASTRUCTURES_IDS = [
            960,
            891,
            182,
            18,
            1083,
            10,
            889,
        ]
        country_codes = self.request.POST.getlist("countries")
        self.country_codes = None
        if "all" not in country_codes:
            self.country_codes = country_codes
        filename = self.generate_filename("Data_networks_report", "xlsx")
        return self.generate_excel(filename)


class EntriesCountReportView(ProtectedTemplateView, EntriesCountReportExcelMixin):
    template_name = "reports/entries_count_report.html"
    permission_classes = (IsAuthenticated,)
    permission_denied_redirect = reverse_lazy("auth:login")

    def get_context_data(self, **kwargs):
        context = super(EntriesCountReportView, self).get_context_data(**kwargs)
        context["form"] = EntriesCountReportForm()
        return context

    def post(self, request, *args, **kwargs):
        self.entrusted_entities_ids = self.request.POST.getlist("entrusted_entities")
        filename = self.generate_filename("Entries_Count_Report", "xlsx")
        return self.generate_excel(filename)


class UserActionsReportView(ProtectedTemplateView, UserActionsReportMixin):
    template_name = "reports/user_actions_report.html"
    permission_classes = (IsAuthenticated,)
    permission_denied_redirect = reverse_lazy("auth:login")

    def get_context_data(self, **kwargs):
        context = super(UserActionsReportView, self).get_context_data(**kwargs)
        context["form"] = UserActionsForm()
        return context

    def post(self, request, *args, **kwargs):
        form = UserActionsForm(request.POST)
        if form.is_valid():
            filename = self.generate_filename("CIS2_User_Actions_Report", "xlsx")
            return self.generate_excel(filename, form.cleaned_data)
        return HttpResponse("Incorrect value selected")
