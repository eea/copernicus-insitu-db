# -*- coding: utf-8 -*-
from __future__ import unicode_literals


from django.apps import apps
from django.core.management import call_command
from django.db import transaction, connection
from django.db.models.fields.reverse_related import ForeignObjectRel
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from insitu import models as insitu_models
from insitu.views import protected
from picklists import models

import os
from io import StringIO

PICKLISTS = [
    insitu_models.CopernicusService, insitu_models.EntrustedEntity,
    insitu_models.Component, models.Barrier, models.ComplianceLevel,
    models.Area, models.Criticality, models.Country, models.DataFormat,
    models.DataType, models.DefinitionLevel, models.Dissemination,
    models.EssentialVariable, models.UpdateFrequency, models.InspireTheme,
    models.ProductGroup, models.ProductStatus, models.Relevance,
    models.RequirementGroup, models.ProviderType,
    models.QualityControlProcedure, models.Timeliness, models.DataPolicy
]

SKIP_FIELDS = ['created_at', 'updated_at']

RELATED_FIELDS = {'entrusted_entity': insitu_models.EntrustedEntity,
                  'service': insitu_models.CopernicusService}


def solve_sql():
    os.environ['DJANGO_COLORS'] = 'nocolor'
    commands = StringIO()
    cursor = connection.cursor()

    for app in apps.get_app_configs():
        label = app.label
        call_command('sqlsequencereset', label, stdout=commands)

    cursor.execute(commands.getvalue())


class ExportPicklistsView(protected.ProtectedView):
    permission_classes = (protected.IsSuperuser,)

    def get(self, request, *args, **kwargs):
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="picklists.xlsx"'
        wb = Workbook()
        wb.remove(wb.active)

        for model in PICKLISTS:
            # Create sheet
            ws = wb.create_sheet(model._meta.model_name)

            # Header
            current_row = 1
            columns = [field for field in model._meta.get_fields()
                       if not isinstance(field, ForeignObjectRel) and field.name not
                       in SKIP_FIELDS]
            for col in columns:
                cell = ws.cell(row=current_row,
                               column=columns.index(col) + 1,
                               value=col.name)
                cell.font = Font(bold=True)

            # Data
            data = model.objects.all()
            for entry in data:
                current_row += 1
                for col in columns:
                    value = getattr(entry, col.name)
                    if col.related_model:
                        value = value.pk
                    ws.cell(row=current_row,
                            column=columns.index(col) + 1,
                            value=value)
            # if 'id' == columns[0]:
            #     ws.column_dimensions.group(start='A',
            #                                end='A',
            #                                hidden=True)
            ws.freeze_panes = 'A2'
        wb.save(response)
        return response


class ImportPicklistsView(protected.ProtectedView):
    permission_classes = (protected.IsSuperuser,)

    def post(self, request, *args, **kwargs):
        if not 'workbook' in request.FILES:
            return HttpResponse(status=400)
        workbook_file = request.FILES['workbook']
        try:
            wb = load_workbook(workbook_file)
            models = {model._meta.model_name: model for model in PICKLISTS}
            name_of_sheet = ""
            for sheet_name in wb.get_sheet_names():
                name_of_sheet = sheet_name
                with transaction.atomic():
                    ws = wb.get_sheet_by_name(sheet_name)
                    model = models.get(sheet_name)
                    fields = [
                        col[0].value
                        for col in ws.iter_cols(min_row=1, max_row=1)
                        if col[0].value is not None
                    ]
                    for row in ws.iter_rows(min_row=2):
                        data = {}
                        pk = row[0].value
                        for i in range(1, len(fields)):
                            field = fields[i]
                            value = row[i].value if row[i].value is not None else ''
                            if field in RELATED_FIELDS.keys():
                                value = RELATED_FIELDS[field].objects.get(pk=value)
                            data[field] = value
                        if not [val for val in data.values() if val]:
                            continue
                        model.objects.update_or_create(pk=pk, defaults=data)
            solve_sql()
        except BaseException as e:
            print(e, " at sheet: ", name_of_sheet)
            return HttpResponse(status=400)
        return HttpResponse(status=200)
